from config import load_app_config
load_app_config()
import sys
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QHBoxLayout, QWidget, QPushButton, QLabel, QComboBox, QListWidget, QTabWidget, QLineEdit, QSpacerItem, QSizePolicy, QListWidgetItem, QFileDialog, QMessageBox, QListWidget, QAbstractItemView, QTableView, QHeaderView, QDialog, QVBoxLayout, QCheckBox, QDialogButtonBox, QLabel, QWidgetAction, QSplitter
)
from PyQt5.QtGui import QFont, QIcon, QStandardItemModel, QStandardItem
from PyQt5.QtCore import Qt, QSortFilterProxyModel, QModelIndex
import pandas as pd
from formulas import calculate_workload_for_employee, get_group_data
from sqlalchemy import and_
from sqlalchemy.orm import sessionmaker, Session
from database import engine
from models import Employee, GroupInstructor, ThesisSupervisors, Reviewer, IndividualRates, OrganizationalUnits, CommitteeFunctionPensum, DidacticCycles, Group, Person, Position, Employment, DidacticCycleClasses, SubjectCycle, Title
from login import LoginWindow
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from PyQt5 import QtCore
from PyQt5.QtCore import Qt
from datetime import datetime
from style.style import light_stylesheet, dark_stylesheet
from openpyxl.drawing.image import Image as ExcelImage
from io import BytesIO
from charts import ChartWidget
from filtersProxy import MultiColumnMultiValueFilterProxyModel
from typing import Optional, List, Dict, Any, Union, Set, Tuple
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

class MainWindow(QMainWindow):
    def __init__(self, user_right: int) -> None:
        super().__init__()
        self.user_right = user_right
        self.group_filter_texts: Dict[int, str] = {}         # {column_index: text}
        self.instructor_filter_texts: Dict[int, str] = {}
        self.summary_filter_texts: Dict[int, str] = {}
        self.current_filtered_groups: List[Dict[str, Any]] = []
        self.current_filtered_instructors: List[Dict[str, Any]] = []
        self.current_filtered_summary: List[Dict[str, Any]] = []
        self.synced_employee_ids: Set[int] = set()
        self.changeFlag: bool = False
        self.setWindowTitle("System Rozliczania Obciążeń Dydaktycznych")
        self.setGeometry(100, 100, 1000, 700)
        self.showMaximized()
        self.is_dark_mode: bool = False  # Domyślnie jasny tryb

        # Główne okno
        main_widget = QWidget()
        main_layout = QVBoxLayout(main_widget)
        main_layout.setContentsMargins(10, 10, 10, 10)

        # Pasek nagłówka
        header = QLabel("System Rozliczania Obciążeń Dydaktycznych")
        header.setFont(QFont("Verdana", 20, QFont.Bold))  # Czcionka "Verdana", rozmiar 20
        header.setAlignment(Qt.AlignmentFlag.AlignCenter)
        header.setStyleSheet("background-color: #2c3e50; color: white; padding: 10px;")
        main_layout.addWidget(header)

        # Filtry
        filters_layout = QVBoxLayout()
        filters_layout.setSpacing(10)

        # Filtry: Rok akademicki
        year_layout = QHBoxLayout()
        year_label = QLabel("Rok akademicki:")
        self.year_filter = QComboBox(self)
        self.year_filter.setMinimumHeight(30)
        self.populate_years()
        year_layout.addWidget(year_label)
        year_layout.addWidget(self.year_filter)
        filters_layout.addLayout(year_layout)

        # Filtry: Jednostka organizacyjna
        unit_layout = QHBoxLayout()
        unit_label = QLabel("Jednostka organizacyjna:")
        self.unit_filter = QComboBox(self)
        self.unit_filter.setMinimumHeight(30)
        self.unit_filter.addItem("Wszystkie jednostki")
        self.populate_units()
        unit_layout.addWidget(unit_label)
        unit_layout.addWidget(self.unit_filter)
        filters_layout.addLayout(unit_layout)

        # Filtry: Wykładowca
        employee_layout = QHBoxLayout()
        employee_label = QLabel("Wykładowca:")
        self.employee_filter = QComboBox(self)
        self.employee_filter.setMinimumHeight(30)
        self.employee_filter.addItem("Wszyscy wykładowcy")
        self.filter_instructors()
        employee_layout.addWidget(employee_label)
        employee_layout.addWidget(self.employee_filter)
        filters_layout.addLayout(employee_layout)

        # Przyciski "Filtruj" i "Odśwież"
        buttons_layout = QHBoxLayout()
        self.filter_button = QPushButton("Filtruj")
        self.filter_button.clicked.connect(self.apply_filters)
        buttons_layout.addWidget(self.filter_button)

        self.refresh_button = QPushButton("Odśwież")
        self.refresh_button.clicked.connect(self.refresh_data)
        buttons_layout.addWidget(self.refresh_button)

        filters_layout.addLayout(buttons_layout)
        self.checkbox_layout = QHBoxLayout()
        self.chceckbox = QCheckBox("Synchronizuj filtry")
        self.checkbox_layout.addWidget(self.chceckbox)
        main_layout.addLayout(filters_layout)
        main_layout.addLayout(self.checkbox_layout)

        # Utwórz QTabWidget i dodaj do layoutu
        self.tab_widget = QTabWidget()
        main_layout.addWidget(self.tab_widget)
        # Przycisk trybu ciemnego/jasnego z ikonką w prawym górnym rogu zakładek
        self.theme_toggle_btn = QPushButton("🌜")
        self.theme_toggle_btn.setCheckable(True)
        self.theme_toggle_btn.setFixedSize(80, 64)
        self.theme_toggle_btn.clicked.connect(self.toggle_theme)
        self.tab_widget.setCornerWidget(self.theme_toggle_btn, Qt.Corner.TopRightCorner)

        # Zakładka grup
        self.groups_tab = QWidget()
        self.groups_layout = QVBoxLayout(self.groups_tab)
        self.group_table = QTableView()
        self.group_model = QStandardItemModel()
        self.group_proxy = MultiColumnMultiValueFilterProxyModel()
        self.group_proxy.setSourceModel(self.group_model)
        self.group_proxy.setFilterCaseSensitivity(1)
        self.group_proxy.setFilterKeyColumn(-1)  # -1 = wszystkie kolumny
        self.group_table.setModel(self.group_proxy)
        header = self.group_table.horizontalHeader()
        if header is not None:
            header.setSectionResizeMode(QHeaderView.ResizeToContents)
        self.group_table.setSortingEnabled(True)
        self.groups_layout.addWidget(self.group_table)
        group_label = QLabel("Grupy:")
        self.group_search = QLineEdit()
        self.group_search.setPlaceholderText("Szukaj w grupach...")
        self.group_search.textChanged.connect(self.filter_group_list)
        self.group_filter_column_combo = QComboBox()
        self.group_filter_column_combo.setMinimumHeight(30)
        self.group_filter_column_combo.currentIndexChanged.connect(self.on_group_filter_column_changed)
        # Dodaj przycisk Wyczyść filtr
        self.clear_group_filter_button = QPushButton("Wyczyść filtry")
        self.clear_group_filter_button.clicked.connect(self.clear_group_filters)
        group_search_layout = QHBoxLayout()
        group_search_layout.addWidget(self.group_filter_column_combo)
        group_search_layout.addWidget(self.group_search)
        group_search_layout.addWidget(self.clear_group_filter_button)
        self.group_active_filters_widget = QWidget()
        self.group_active_filters_layout = QVBoxLayout(self.group_active_filters_widget)
        self.groups_layout.addWidget(self.group_active_filters_widget)
        self.groups_layout.addLayout(group_search_layout)
        self.groups_layout.addLayout(group_search_layout)
        self.groups_layout.addWidget(self.group_table)
        self.tab_widget.addTab(self.groups_tab, "Grupy")

        # Zakładka wykładowców
        self.instructors_tab = QWidget()
        self.instructors_layout = QVBoxLayout(self.instructors_tab)
        self.instructor_search = QLineEdit()
        self.instructor_search.setPlaceholderText("Szukaj w wykładowcach...")
        self.instructor_search.textChanged.connect(self.filter_instructor_list)
        self.instructor_active_filters_widget = QWidget()
        self.instructor_active_filters_layout = QVBoxLayout(self.instructor_active_filters_widget)
        self.instructors_layout.addWidget(self.instructor_active_filters_widget)
        self.instructor_filter_column_combo = QComboBox()
        self.instructor_filter_column_combo.setMinimumHeight(60)
        self.instructor_filter_column_combo.setMinimumWidth(120)
        self.instructor_filter_column_combo.currentIndexChanged.connect(self.on_instructor_filter_column_changed)
        # Dodaj przycisk Wyczyść filtr
        self.clear_instructor_filter_button = QPushButton("Wyczyść filtry")
        self.clear_instructor_filter_button.clicked.connect(self.clear_instructor_filters)
        instructor_search_layout = QHBoxLayout()
        instructor_search_layout.addWidget(self.instructor_filter_column_combo)
        instructor_search_layout.addWidget(self.instructor_search)
        instructor_search_layout.addWidget(self.clear_instructor_filter_button)
        self.instructors_layout.addLayout(instructor_search_layout)
        self.instructor_table = QTableView()
        self.instructor_model = QStandardItemModel()
        self.instructor_proxy = MultiColumnMultiValueFilterProxyModel()
        self.instructor_proxy.setSourceModel(self.instructor_model)
        self.instructor_proxy.setFilterCaseSensitivity(1)
        self.instructor_proxy.setFilterKeyColumn(-1)
        self.instructor_table.setModel(self.instructor_proxy)
        header = self.instructor_table.horizontalHeader()
        if header is not None:
            header.setSectionResizeMode(QHeaderView.ResizeToContents)
        self.instructor_table.setSortingEnabled(True)
        self.instructors_layout.addWidget(self.instructor_table)
        details_label = QLabel("Szczegóły wykładowcy:")
        self.instructor_details_table = QTableView()
        self.instructor_details_model = QStandardItemModel()
        self.instructor_details_table.setModel(self.instructor_details_model)
        header = self.instructor_details_table.horizontalHeader()
        if header is not None:
            header.setSectionResizeMode(QHeaderView.ResizeToContents)
        self.instructors_layout.addWidget(self.instructor_details_table)
        self.tab_widget.addTab(self.instructors_tab, "Wykładowcy")
        self.instructor_table.clicked.connect(self.display_instructor_details)

        # Zakładka zestawienie z wykresami
        self.setup_summary_tab_with_charts()

        # Przycisk "Generuj raport" na samym dole
        self.report_button = QPushButton("Generuj raport")
        self.report_button.clicked.connect(self.generate_report)
        main_layout.addWidget(self.report_button)

        # Status bar
        self.status_label = QLabel("Status: Oczekiwanie na akcję")
        main_layout.addWidget(self.status_label)

        self.setCentralWidget(main_widget)
        # Populate initial data
        self.populate_groups()
        self.populate_employees()
        self.populate_summary()
        self.year_filter.currentIndexChanged.connect(self.filter_instructors)
        self.unit_filter.currentIndexChanged.connect(self.filter_instructors)
        self.year_filter.setToolTip("Wybierz rok akademicki")
        self.unit_filter.setToolTip("Wybierz jednostkę organizacyjną")
        self.employee_filter.setToolTip("Wybierz wykładowcę")
        self.filter_button.setToolTip("Zastosuj wybrane filtry do danych")
        self.refresh_button.setToolTip("Odśwież dane bez zmiany filtrów")
        self.group_search.setToolTip("Wyszukaj grupę po dowolnym polu")
        self.clear_group_filter_button.setToolTip("Wyczyść pole wyszukiwania grup")
        self.instructor_search.setToolTip("Wyszukaj wykładowcę po nazwisku lub imieniu")
        self.clear_instructor_filter_button.setToolTip("Wyczyść pole wyszukiwania wykładowców")
        self.report_button.setToolTip("Wygeneruj raport Excel z aktualnych danych")
        self.theme_toggle_btn = QPushButton("🌜")
        self.theme_toggle_btn.setObjectName("ThemeToggle")
        self.theme_toggle_btn.setFixedSize(80, 64)
        self.theme_toggle_btn.setCheckable(True)
        self.theme_toggle_btn.clicked.connect(self.toggle_theme)
        self.tab_widget.setCornerWidget(self.theme_toggle_btn, Qt.Corner.TopRightCorner)
        self.theme_toggle_btn.setToolTip("Przełącz tryb ciemny/jasny")
        if hasattr(self, 'theme_toggle_btn'):
            self.theme_toggle_btn.setToolTip("Przełącz tryb ciemny/jasny")
        self.setStyleSheet(light_stylesheet)
        self.tab_widget.currentChanged.connect(self.refresh_data)
        self.chceckbox.stateChanged.connect(self.toogle_checkbox)

    def setup_summary_tab_with_charts(self) -> None:
        """Metoda dla zakładki zestawienia z wykresami w osobnych zakładkach"""
        self.summary_tab = QWidget()
        self.summary_layout = QVBoxLayout(self.summary_tab)

        # Utwórz QTabWidget do przechowywania zakładek
        self.summary_tab_widget = QTabWidget()
        
        # ZAKŁADKA 1 - TABELA
        table_container = QWidget()
        table_layout = QVBoxLayout(table_container)
        
        # Filtry dla tabeli
        filter_container = QWidget()
        filter_layout = QVBoxLayout(filter_container)
        
        self.summary_search = QLineEdit()
        self.summary_search.setPlaceholderText("Szukaj w podsumowaniu...")
        self.summary_search.textChanged.connect(self.filter_summary_list)
        self.summary_active_filters_widget = QWidget()
        self.summary_active_filters_layout = QVBoxLayout(self.summary_active_filters_widget)
        self.summary_layout.addWidget(self.summary_active_filters_widget)
        self.summary_filter_column_combo = QComboBox()
        self.summary_filter_column_combo.setMinimumHeight(30)
        self.summary_filter_column_combo.currentIndexChanged.connect(self.on_summary_filter_column_changed)
        
        self.clear_summary_filter_button = QPushButton("Wyczyść filtr")
        self.clear_summary_filter_button.clicked.connect(self.clear_summary_filters)
        
        search_row = QHBoxLayout()
        search_row.addWidget(self.summary_filter_column_combo)
        search_row.addWidget(self.summary_search)
        search_row.addWidget(self.clear_summary_filter_button)
        
        filter_layout.addLayout(search_row)
        table_layout.addWidget(filter_container)
        
        # Tabela podsumowania
        self.summary_table = QTableView()
        self.summary_model = QStandardItemModel()
        self.summary_proxy = MultiColumnMultiValueFilterProxyModel()
        self.summary_proxy.setSourceModel(self.summary_model)
        self.summary_proxy.setFilterCaseSensitivity(1)
        self.summary_proxy.setFilterKeyColumn(-1)
        self.summary_table.setModel(self.summary_proxy)
        self.summary_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.summary_table.setSortingEnabled(True)
        
        table_layout.addWidget(self.summary_table)
        table_container.setLayout(table_layout)
        
        # Dodaj zakładkę z tabelą
        self.summary_tab_widget.addTab(table_container, "Tabela")

        # ZAKŁADKA 2 - WYKRESY
        self.chart_widget = ChartWidget()
        self.chart_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        
        # Dodaj zakładkę z wykresami
        self.summary_tab_widget.addTab(self.chart_widget, "Wykresy")
        
        # Dodaj QTabWidget do layoutu
        self.summary_layout.addWidget(self.summary_tab_widget)

        # Tooltips
        self.summary_search.setToolTip("Wyszukaj w podsumowaniu po dowolnym polu")
        self.clear_summary_filter_button.setToolTip("Wyczyść pole wyszukiwania w podsumowaniu")

        self.tab_widget.addTab(self.summary_tab, "Zestawienia")

    def toogle_checkbox(self) -> None:
        """Toggle the checkbox state."""
        self.changeFlag = True
        if not self.chceckbox.isChecked():
            self.synced_employee_ids.clear()
        print("Checkbox został zmieniony, odświeżam dane.")
        self.refresh_data()    

    def refresh_data(self) -> None:
        if self.changeFlag:
            if self.chceckbox.isChecked():
                # Dodaj ID z aktualnego widoku
                current_ids = set(self.get_instructors_id(self.instructor_proxy))
                self.synced_employee_ids.update(current_ids)
                filtered_employee_ids: Optional[List[int]] = list(self.synced_employee_ids)
            else:
                filtered_employee_ids = None

            self.populate_groups(filtered_employee_ids=filtered_employee_ids)
            self.populate_employees()
            self.populate_summary()
            self.changeFlag = False
            self.status_label.setText("Dane odświeżone.")

    def on_tab_changed(self, index: int) -> None:
        """Handle tab change events."""
        if self.tab_widget.tabText(index) == "Wykładowcy":
            self.populate_employees()

    def populate_years(self) -> None:
        """Populate the year filter with distinct academic years from DidacticCycles."""
        self.year_filter.clear()
        db: Session = SessionLocal()
        try:
            # Pobierz unikalne lata akademickie
            years: List[tuple[str]] = db.query(DidacticCycles.OPIS).join(SubjectCycle, SubjectCycle.CDYD_KOD==DidacticCycles.KOD).distinct().all()

            # Wyodrębnij fragment "2024/25" z "Rok akademicki 2024/25"
            unique_years: List[str] = sorted(list(set(year[0].split()[-1] for year in years if year[0] is not None)),
                                reverse=True)  # Sortuj malejąco

            # Dodaj każdy rok akademicki do filtra
            for year in unique_years:
                self.year_filter.addItem(year)
        except Exception as e:
            self.status_label.setText(f"Status: Błąd podczas pobierania lat: {str(e)}")
            print(f"Database error details: {str(e)}")
        finally:
            db.close()
    
    def populate_units(self) -> None:
        """Populate the unit filter with only allowed institutes based on user rights."""
        self.unit_filter.clear()
        db: Session = SessionLocal()
        print(f"User right: {self.user_right}")  # Debugging: Log the user right
        try:
            # Pobierz jednostki organizacyjne na podstawie prawa użytkownika
            units: List[OrganizationalUnits]
            if self.user_right == 0:  # Dostęp do wszystkich jednostek
                self.unit_filter.addItem("Wszystkie jednostki", None) 
                units = db.query(OrganizationalUnits).filter(OrganizationalUnits.OPIS.like("Instytut %")).all()
            elif self.user_right == 1:  # Dostęp tylko do Instytutu Informatyki Stosowanej
                units = db.query(OrganizationalUnits).filter(OrganizationalUnits.OPIS == "Instytut Informatyki Stosowanej im. Krzysztofa Brzeskiego").all()
            elif self.user_right == 4:  # Dostęp tylko do Instytutu Ekonomicznego
                units = db.query(OrganizationalUnits).filter(OrganizationalUnits.OPIS == "Instytut Ekonomiczny").all()
            elif self.user_right == 3:  # Dostęp tylko do Instytutu Politechnicznego
                units = db.query(OrganizationalUnits).filter(OrganizationalUnits.OPIS == "Instytut Politechniczny").all()
            elif self.user_right == 2:  # Dostęp tylko do Instytutu Pedagogiczno-Językowego
                units = db.query(OrganizationalUnits).filter(OrganizationalUnits.OPIS == "Instytut Pedagogiczno- Językowy").all()
            else:
                units = []  # Brak dostępu do żadnych jednostek
                self.status_label.setText("Błąd: Nieprawidłowa rola użytkownika. Skontaktuj się z administratorem.")
                QMessageBox.critical(self, "Błąd uprawnień", "Nieprawidłowa rola użytkownika. Skontaktuj się z administratorem.")

            # Dodaj jednostki do filtra
            for unit in units:
                self.unit_filter.addItem(str(unit.OPIS), unit.KOD)
        except Exception as e:
            print(f"Błąd podczas ładowania jednostek organizacyjnych: {str(e)}")
        finally:
            db.close()
    
    def apply_filters(self) -> None:
        """Apply filters and refresh data in both tabs."""
        self.populate_groups()
        self.filter_instructors() 
        self.populate_employees()
        self.populate_summary()
        self.status_label.setText("Status: Filtry zostały zastosowane.")

    def filter_instructors(self) -> None:
        """Filter and populate the instructor list based on the selected unit."""
        db: Session = SessionLocal()
        selected_unit: Optional[int] = self.unit_filter.currentData()
        selected_year: str = self.year_filter.currentText()

        current_instructor: Optional[int] = self.employee_filter.currentData()

        try:
            # Query instructors based on the selected unit
            instructor_query = (
                db.query(Employee)
                .join(Person, Employee.OS_ID == Person.ID)
                .join(GroupInstructor, Employee.ID == GroupInstructor.PRAC_ID)
                .join(OrganizationalUnits, GroupInstructor.JEDN_KOD == OrganizationalUnits.KOD)
                .join(Group, GroupInstructor.ZAJ_CYK_ID == Group.ZAJ_CYK_ID)
                .join(DidacticCycleClasses, Group.ZAJ_CYK_ID == DidacticCycleClasses.ID)
                .join(DidacticCycles, DidacticCycleClasses.CDYD_KOD == DidacticCycles.KOD)
            )
            if selected_unit:
                instructor_query = instructor_query.filter(GroupInstructor.JEDN_KOD == selected_unit)
            if selected_year:
                instructor_query = instructor_query.filter(DidacticCycles.OPIS.like(f"%{selected_year}%"))
            instructors: List[Employee] = instructor_query.all()
            instructors.sort(key=lambda i: (getattr(i.Person, 'NAZWISKO', ''), getattr(i.Person, 'IMIE', '')) if hasattr(i, 'Person') and i.Person else (getattr(db.query(Person).filter_by(ID=i.OS_ID).first(), 'NAZWISKO', ''), getattr(db.query(Person).filter_by(ID=i.OS_ID).first(), 'IMIE', '')))
            self.employee_filter.clear()
            self.employee_filter.addItem("Wszyscy wykładowcy", None)
            for instructor in instructors:
                person: Optional[Person] = db.query(Person).filter_by(ID=instructor.OS_ID).first()
                self.employee_filter.addItem(f"{getattr(person, 'NAZWISKO', 'Brak')} {getattr(person, 'IMIE', 'Brak')}", instructor.ID)
            
            index_to_restore: int = self.employee_filter.findData(current_instructor)
            if index_to_restore != -1:
                self.employee_filter.setCurrentIndex(index_to_restore)
        except Exception as e:
            print(f"Error: {str(e)}")  # Debugging: Log the error
        finally:
            db.close()

    def display_instructor_details(self, index: QModelIndex) -> None:
        self.instructor_details_model.clear()
        source_index: QModelIndex = self.instructor_proxy.mapToSource(index)
        row: int = source_index.row()
        item: Optional[QStandardItem] = self.instructor_model.item(row, 1)
        if item is not None:
            nazwisko_imie: str = item.text()
        else:
            nazwisko_imie = ""
        selected_year: str = self.year_filter.currentText()
        selected_unit: Optional[int] = self.unit_filter.currentData()
        db: Session = SessionLocal()
        try:
            person: Optional[Person] = db.query(Person).filter(
                (Person.NAZWISKO + " " + Person.IMIE) == nazwisko_imie
            ).first()
            if not person:
                self.instructor_details_model.setHorizontalHeaderLabels(["Informacja"])
                self.instructor_details_model.appendRow([QStandardItem("Nie znaleziono wykładowcy.")])
                return
            employee: Optional[Employee] = db.query(Employee).filter(Employee.OS_ID == person.ID).first()
            if not employee:
                self.instructor_details_model.setHorizontalHeaderLabels(["Informacja"])
                self.instructor_details_model.appendRow([QStandardItem("Nie znaleziono pracownika.")])
                return
            selected_employee_id: int = employee.ID

            group_data: List[Dict[str, Any]]
            if self.chceckbox.isChecked():
                group_data = get_group_data(selected_year, selected_unit, selected_employee_id, self.current_filtered_groups)
            else:
                group_data = get_group_data(selected_year, selected_unit, selected_employee_id, None)
            
            filtered_groups: Optional[List[Dict[str, Any]]] = self.current_filtered_groups if self.chceckbox.isChecked() else None
            if filtered_groups:
                group_data = [group for group in group_data if group in filtered_groups]
            
            workload_data: Dict[str, Any] = calculate_workload_for_employee(employee.ID, selected_year, selected_unit, filtered_groups)

            # Dane podstawowe
            headers: List[str] = [
                "Stanowisko", "Pensum uczelniane", "Umowa od", "Umowa do", "Pensum", "Godziny Z stacjonarne", "Godziny Z niestacjonarne",
                "Godziny L stacjonarne", "Godziny L niestacjonarne", "Nadgodziny/Niedobór", "Łączna zniżka", "Etat", "Czy podstawowe miejsce pracy",
                "Stawka", "Kwota nadgodzin"
            ]
            self.instructor_details_model.setHorizontalHeaderLabels(headers)
            row_items: List[QStandardItem] = [
                QStandardItem(str(workload_data.get('stanowisko', ''))),
                QStandardItem(str(workload_data.get('pensum_uczelniane', ''))),
                QStandardItem(str(workload_data.get('umowa_pocz', ''))),
                QStandardItem(str(workload_data.get('umowa_kon', ''))),
                QStandardItem(str(workload_data.get('pensum', ''))),
                QStandardItem(str(workload_data.get('godziny_dydaktyczne_z_stacjonarne', ''))),
                QStandardItem(str(workload_data.get('godziny_dydaktyczne_z_niestacjonarne', ''))),
                QStandardItem(str(workload_data.get('godziny_dydaktyczne_l_stacjonarne', ''))),
                QStandardItem(str(workload_data.get('godziny_dydaktyczne_l_niestacjonarne', ''))),
                QStandardItem(str(workload_data.get('nadgodziny', ''))),
                QStandardItem(str(workload_data.get('zniżka', ''))),
                QStandardItem(str(workload_data.get('etat', ''))),
                QStandardItem(str(workload_data.get('CZY_PODSTAWOWE', ''))),
                QStandardItem(str(workload_data.get('stawka', ''))),
                QStandardItem(str(workload_data.get('kwota_nadgodzin', '')))
            ]
            self.instructor_details_model.appendRow(row_items)

            # Zniżki
            self.instructor_details_model.appendRow([QStandardItem("--- Zniżki ---")] + [QStandardItem("") for _ in range(len(headers)-1)])
            if workload_data.get("typy_znizek") and workload_data["typy_znizek"] != ["Brak zniżek"]:
                for znizka, godziny in zip(workload_data["typy_znizek"], workload_data.get("godziny_znizek", [])):
                    self.instructor_details_model.appendRow([QStandardItem(f"{znizka} ({godziny} godz.)")] + [QStandardItem("") for _ in range(len(headers)-1)])
            else:
                self.instructor_details_model.appendRow([QStandardItem("Brak zniżek")] + [QStandardItem("") for _ in range(len(headers)-1)])

            # Przedmioty
            self.instructor_details_model.appendRow([QStandardItem("--- Przedmioty ---")] + [QStandardItem("") for _ in range(len(headers)-1)])
            self.instructor_details_model.appendRow([QStandardItem("Przedmiot"), QStandardItem("Typ zajęć"), QStandardItem("Liczba godzin"), QStandardItem("Semestr"), QStandardItem("Kierunek"), QStandardItem("Specjalność"), QStandardItem("Tryb"), QStandardItem("Instytut w którym jest rozliczany przedmiot")] + [QStandardItem("") for _ in range(len(headers)-7)])
            for group in group_data:
                self.instructor_details_model.appendRow([
                    QStandardItem(group.get('Przedmiot', '')),
                    QStandardItem(group.get('Typ zajęć', '')),
                    QStandardItem(str(group.get('Liczba godzin', ''))),
                    QStandardItem(group.get('Semestr', '')),
                    QStandardItem(f"{group.get('Kierunek', '')}"),
                    QStandardItem(f"{group.get('Specjalność', '')}"),
                    QStandardItem(group.get('Tryb', '')),
                    QStandardItem(group.get('Instytut w którym jest rozliczany przedmiot', '')),
                ] + [QStandardItem("") for _ in range(len(headers)-7)])

            # Przedmioty (podsumowanie)
            self.instructor_details_model.appendRow([QStandardItem("--- Przedmioty (podsumowanie) ---")] + [QStandardItem("") for _ in range(len(headers)-1)])
            
            # Nagłówki dla podsumowania przedmiotów
            subject_summary_headers: List[str] = [
                "Kierunek", "Specjalność",
                "Zimowy stacjonarne", "Zimowy niestacjonarne",
                "Letni stacjonarne", "Letni niestacjonarne",
                "Suma"
            ]
            self.instructor_details_model.appendRow([QStandardItem(h) for h in subject_summary_headers] + [QStandardItem("") for _ in range(len(headers) - len(subject_summary_headers))])

            kierunek_dict: Dict[str, Dict[str, Dict[str, Union[int, float]]]] = {}
            specjalnosc_display_names: Dict[tuple[str, str], str] = {}

            for group in group_data:
                kierunek: str = group.get("Kierunek", "Nieznany kierunek")
                specjalnosc: str = group.get("Specjalność", "Brak specjalności")
                specjalnosc_key: str = specjalnosc.strip().lower()
                specjalnosc_display_names[(kierunek, specjalnosc_key)] = specjalnosc
                tryb: str = group.get("Tryb", "Nieznany tryb").strip().lower()
                hours: Union[int, float] = group.get("Liczba godzin", 0)
                semester: str = group.get("Semestr", "Nieznany semestr").lower()

                if kierunek not in kierunek_dict:
                    kierunek_dict[kierunek] = {}
                if specjalnosc_key not in kierunek_dict[kierunek]:
                    kierunek_dict[kierunek][specjalnosc_key] = {
                        "Zimowy stacjonarne": 0,
                        "Zimowy niestacjonarne": 0,
                        "Letni stacjonarne": 0,
                        "Letni niestacjonarne": 0,
                        "Suma": 0
                    }
                if "zimowy" in semester and (("niestacjonarne" in tryb) or tryb == "none"):
                    kierunek_dict[kierunek][specjalnosc_key]["Zimowy niestacjonarne"] += hours
                elif "zimowy" in semester and "stacjonarne" in tryb:
                    kierunek_dict[kierunek][specjalnosc_key]["Zimowy stacjonarne"] += hours
                elif "letni" in semester and (("niestacjonarne" in tryb) or tryb == "none"):
                    kierunek_dict[kierunek][specjalnosc_key]["Letni niestacjonarne"] += hours
                elif "letni" in semester and "stacjonarne" in tryb:
                    kierunek_dict[kierunek][specjalnosc_key]["Letni stacjonarne"] += hours
                else:
                    print(f"Nieznany tryb/semestr w display_instructor_details: tryb={tryb}, semester={semester}, hours={hours}, kierunek={kierunek}, specjalnosc={specjalnosc}")

                kierunek_dict[kierunek][specjalnosc_key]["Suma"] += hours

            for kierunek, specjalnosci in kierunek_dict.items():
                suma_kierunku: Dict[str, Union[int, float]] = {
                    "Zimowy stacjonarne": 0,
                    "Zimowy niestacjonarne": 0,
                    "Letni stacjonarne": 0,
                    "Letni niestacjonarne": 0,
                    "Suma": 0
                }
                for specjalnosc_key, godziny in specjalnosci.items():
                    specjalnosc: str = specjalnosc_display_names.get((kierunek, specjalnosc_key), specjalnosc_key)
                    row_items_subject: List[QStandardItem] = [
                        QStandardItem(kierunek),
                        QStandardItem(specjalnosc),
                        QStandardItem(str(godziny["Zimowy stacjonarne"])),
                        QStandardItem(str(godziny["Zimowy niestacjonarne"])),
                        QStandardItem(str(godziny["Letni stacjonarne"])),
                        QStandardItem(str(godziny["Letni niestacjonarne"])),
                        QStandardItem(str(godziny["Suma"]))
                    ]
                    self.instructor_details_model.appendRow(row_items_subject + [QStandardItem("") for _ in range(len(headers) - len(subject_summary_headers))])
                    
                    # Dodaj do sum kierunku
                    suma_kierunku["Zimowy stacjonarne"] += godziny["Zimowy stacjonarne"]
                    suma_kierunku["Zimowy niestacjonarne"] += godziny["Zimowy niestacjonarne"]
                    suma_kierunku["Letni stacjonarne"] += godziny["Letni stacjonarne"]
                    suma_kierunku["Letni niestacjonarne"] += godziny["Letni niestacjonarne"]
                    suma_kierunku["Suma"] += godziny["Suma"]
                
                # Dodaj wiersz sumujący dla kierunku
                row_items_sum_kierunek: List[QStandardItem] = [
                    QStandardItem(kierunek),
                    QStandardItem("SUMA kierunku"),
                    QStandardItem(str(suma_kierunku["Zimowy stacjonarne"])),
                    QStandardItem(str(suma_kierunku["Zimowy niestacjonarne"])),
                    QStandardItem(str(suma_kierunku["Letni stacjonarne"])),
                    QStandardItem(str(suma_kierunku["Letni niestacjonarne"])),
                    QStandardItem(str(suma_kierunku["Suma"]))
                ]
                self.instructor_details_model.appendRow(row_items_sum_kierunek + [QStandardItem("") for _ in range(len(headers) - len(subject_summary_headers))])

            if not kierunek_dict:
                self.instructor_details_model.appendRow([QStandardItem("Brak przedmiotów do wyświetlenia.")] + [QStandardItem("") for _ in range(len(headers)-1)])

        except Exception as e:
            self.instructor_details_model.setHorizontalHeaderLabels(["Błąd"])
            self.instructor_details_model.appendRow([QStandardItem(f"Błąd: {str(e)}")])
        finally:
            db.close()


    def populate_groups(self, filtered_employee_ids: Optional[List[int]] = None) -> None:
        self.group_model.clear()
        selected_unit: Optional[int] = self.unit_filter.currentData()
        selected_year: str = self.year_filter.currentText()
        selected_employee: Optional[int] = self.employee_filter.currentData()

        try:     
            group_data: List[Dict[str, Any]]
            if self.chceckbox.isChecked():
                # W trybie synchronizacji pobieramy pełne dane (bez zawężania po current_filtered_groups),
                # a filtrowanie widoku robi proxy na podstawie bieżących tekstów filtrów.
                group_data = get_group_data(selected_year, selected_unit, None, None, filtered_employee_ids)
            else:
                group_data = get_group_data(selected_year, selected_unit, selected_employee, None)

            if not group_data:
                self.group_model.setHorizontalHeaderLabels(["Brak danych do wyświetlenia."])
                return
            # Ustal nagłówki na podstawie kluczy pierwszego rekordu
            headers: List[str] = list(group_data[0].keys())
            self.group_model.setHorizontalHeaderLabels(headers)
            self.update_group_filter_columns(headers)
            for group in group_data:
                row_items: List[QStandardItem] = []
                for col in headers:
                    value: Any = group.get(col, "")
                    item = QStandardItem(str(value))  # <-- Tworzymy nowy obiekt za każdym razem
                    # Jeśli wartość jest liczbą (int lub float), ustaw dane liczbowe
                    try:
                        num: float = float(value)
                        item.setData(num, 2)
                    except (ValueError, TypeError):
                        pass  # zostaw jako tekst
                    row_items.append(item)
                self.group_model.appendRow(row_items)
            self.save_current_filtered_groups()
        except Exception as e:
            self.group_model.setHorizontalHeaderLabels(["Błąd"])
            self.group_model.appendRow([QStandardItem(f"Błąd: {str(e)}")])
        
    def populate_employees(self) -> None:
        self.instructor_model.clear()
        selected_unit: Optional[int] = self.unit_filter.currentData()
        selected_year: str = self.year_filter.currentText()
        selected_employee: Optional[int] = self.employee_filter.currentData()
        db: Session = SessionLocal()
        try:
            query = db.query(Employee, Person).join(Person, Employee.OS_ID == Person.ID).filter(GroupInstructor.PRAC_ID == Employee.ID).filter(DidacticCycles.OPIS.like(f"%{selected_year}%"))
            if selected_unit:
                query = query.filter(GroupInstructor.JEDN_KOD == selected_unit)
            if selected_employee:
                query = query.filter(Employee.ID == selected_employee)
            results: List[tuple[Employee, Person]] = query.all()
            results.sort(key=lambda pair: (pair[1].NAZWISKO, pair[1].IMIE))
            headers: List[str] = [
                "Tytuły", "Nazwisko i imię", "J.O.", "Forma", "Stanowisko", "Umowa od", "Umowa do",
                "Pensum uczelniane", "Zniżka", "Czy podstawowe miejsce pracy", "Godziny dydaktyczne Z stacjonarne",
                "Godziny dydaktyczne Z niestacjonarne", "Godziny dydaktyczne L stacjonarne", "Godziny dydaktyczne L niestacjonarne", "Pensum realne", "Pensum", "Etat", "Nadgodziny", "Stawka", "Kwota nadgodzin"
            ]
            self.instructor_model.setHorizontalHeaderLabels(headers)
            self.update_instructor_filter_columns(headers)
            for employee, person in results:
                filtered_groups: Optional[List[Dict[str, Any]]] = self.current_filtered_groups if self.chceckbox.isChecked() else None
                workload_data: Dict[str, Any] = calculate_workload_for_employee(employee.ID, selected_year, selected_unit, filtered_groups)
                if workload_data["total_workload"] > 0:
                    db2: Session = SessionLocal()
                    tytul: Optional[Title] = db2.query(Title).filter_by(ID=person.TYTUL_PRZED).first()
                    organizational_unit: Optional[OrganizationalUnits] = db2.query(OrganizationalUnits).filter_by(KOD=person.JED_ORG_KOD).first()
                    db2.close()
                    # Tworzymy QStandardItemy
                    tytul_str: str = str(tytul.NAZWA) if tytul and hasattr(tytul, 'NAZWA') else "N/A"
                    nazwisko: str = getattr(person, 'NAZWISKO', 'Brak') if person else 'Brak'
                    imie: str = getattr(person, 'IMIE', 'Brak') if person else 'Brak'
                    organizational_unit_str: str = str(organizational_unit.OPIS) if organizational_unit and hasattr(organizational_unit, 'OPIS') else "N/A"
                    row_items: List[QStandardItem] = [
                        QStandardItem(tytul_str),
                        QStandardItem(f"{nazwisko} {imie}"),
                        QStandardItem(organizational_unit_str),
                        QStandardItem("etat" if workload_data["umowa_pocz"] != "Brak daty rozpoczęcia umowy" else "umowa zlecenie"),
                        QStandardItem(str(workload_data['stanowisko'])),
                        QStandardItem(str(workload_data['umowa_pocz'])),
                        QStandardItem(str(workload_data['umowa_kon'])),
                        QStandardItem(str(workload_data['pensum_uczelniane'])),
                        QStandardItem(str(workload_data['zniżka'])),
                        QStandardItem(str(workload_data['CZY_PODSTAWOWE'])),
                        QStandardItem(str(workload_data['godziny_dydaktyczne_z_stacjonarne'])),
                        QStandardItem(str(workload_data['godziny_dydaktyczne_z_niestacjonarne'])),
                        QStandardItem(str(workload_data['godziny_dydaktyczne_l_stacjonarne'])),
                        QStandardItem(str(workload_data['godziny_dydaktyczne_l_niestacjonarne'])),
                        QStandardItem(str(workload_data['total_workload'])),
                        QStandardItem(str(workload_data['pensum'])),
                        QStandardItem(str(workload_data['etat'])),
                        QStandardItem(str(workload_data['nadgodziny'])),
                        QStandardItem(str(workload_data['stawka'])),
                        QStandardItem(str(workload_data['kwota_nadgodzin'])),
                    ]
                    # Ustaw dane liczbowe dla kolumn liczbowych
                    numeric_indices: List[int] = [7,8,10,11,12,13,14,15,16,17]  # indeksy kolumn liczbowych
                    numeric_keys: List[str] = [
                        'pensum_uczelniane','zniżka','godziny_dydaktyczne_z_stacjonarne','godziny_dydaktyczne_z_niestacjonarne','godziny_dydaktyczne_l_stacjonarne','godziny_dydaktyczne_l_niestacjonarne',
                        'total_workload','pensum','etat','nadgodziny','stawka','kwota_nadgodzin'
                    ]
                    for idx, key in zip(numeric_indices, numeric_keys):
                        try:
                            value: float = float(workload_data[key])
                            row_items[idx].setData(value, 2)
                        except (ValueError, TypeError):
                            pass
                    self.instructor_model.appendRow(row_items)
            if not results:
                self.instructor_model.setHorizontalHeaderLabels(["Brak wykładowców do wyświetlenia."])
            self.save_current_filtered_instructors()
        except Exception as e:
            self.instructor_model.setHorizontalHeaderLabels(["Błąd"])
            self.instructor_model.appendRow([QStandardItem(f"Błąd: {str(e)}")])
        finally:
            db.close()

    def populate_summary(self) -> None:
        self.summary_model.clear()
        selected_unit: Optional[int] = self.unit_filter.currentData()
        selected_year: str = self.year_filter.currentText()
        selected_employee: Optional[int] = self.employee_filter.currentData()

        try:     
            group_data: List[Dict[str, Any]]
            if self.chceckbox.isChecked():
                group_data = get_group_data(selected_year, selected_unit, selected_employee, self.current_filtered_groups)
            else:
                group_data = get_group_data(selected_year, selected_unit, selected_employee, None)

            kierunek_dict: Dict[str, Dict[str, Dict[str, Union[int, float]]]] = {}
            specjalnosc_display_names: Dict[tuple[str, str], str] = {}

            for group in group_data:
                kierunek: str = group.get("Kierunek", "Nieznany kierunek")
                specjalnosc: str = group.get("Specjalność", "Brak specjalności")
                specjalnosc_key: str = specjalnosc.strip().lower()
                specjalnosc_display_names[(kierunek, specjalnosc_key)] = specjalnosc
                tryb: str = group.get("Tryb", "Nieznany tryb").strip().lower()
                hours: Union[int, float] = group.get("Liczba godzin", 0)
                semester: str = group.get("Semestr", "Nieznany semestr").lower()

                if kierunek not in kierunek_dict:
                    kierunek_dict[kierunek] = {}
                if specjalnosc_key not in kierunek_dict[kierunek]:
                    kierunek_dict[kierunek][specjalnosc_key] = {
                        "Zimowy stacjonarne": 0,
                        "Zimowy niestacjonarne": 0,
                        "Letni stacjonarne": 0,
                        "Letni niestacjonarne": 0,
                        "Suma": 0
                    }
                if "zimowy" in semester and (("niestacjonarne" in tryb) or tryb == "none"):
                    kierunek_dict[kierunek][specjalnosc_key]["Zimowy niestacjonarne"] = float(kierunek_dict[kierunek][specjalnosc_key]["Zimowy niestacjonarne"]) + hours
                elif "zimowy" in semester and "stacjonarne" in tryb:
                    kierunek_dict[kierunek][specjalnosc_key]["Zimowy stacjonarne"] = float(kierunek_dict[kierunek][specjalnosc_key]["Zimowy stacjonarne"]) + hours
                elif "letni" in semester and (("niestacjonarne" in tryb) or tryb == "none"):
                    kierunek_dict[kierunek][specjalnosc_key]["Letni niestacjonarne"] = float(kierunek_dict[kierunek][specjalnosc_key]["Letni niestacjonarne"]) + hours
                elif "letni" in semester and "stacjonarne" in tryb:
                    kierunek_dict[kierunek][specjalnosc_key]["Letni stacjonarne"] = float(kierunek_dict[kierunek][specjalnosc_key]["Letni stacjonarne"]) + hours
                else:
                    # Jeśli tryb nie jest rozpoznany, możesz dodać do osobnej kolumny lub wyświetlić ostrzeżenie
                    print(f"Nieznany tryb/semestr: tryb={tryb}, semester={semester}, hours={hours}, kierunek={kierunek}, specjalnosc={specjalnosc}")

                kierunek_dict[kierunek][specjalnosc_key]["Suma"] = float(kierunek_dict[kierunek][specjalnosc_key]["Suma"]) + hours

            headers: List[str] = [
                "Kierunek", "Specjalność",
                "Zimowy stacjonarne", "Zimowy niestacjonarne",
                "Letni stacjonarne", "Letni niestacjonarne",
                "Suma"
            ]
            self.summary_model.setHorizontalHeaderLabels(headers)
            self.update_summary_filter_columns(headers)

            for kierunek, specjalnosci in kierunek_dict.items():
                # Sumy dla kierunku
                suma_kierunku: Dict[str, Union[int, float]] = {
                    "Zimowy stacjonarne": 0,
                    "Zimowy niestacjonarne": 0,
                    "Letni stacjonarne": 0,
                    "Letni niestacjonarne": 0,
                    "Suma": 0
                }
                for specjalnosc_key, godziny in specjalnosci.items():
                    specjalnosc: str = specjalnosc_display_names.get((kierunek, specjalnosc_key), specjalnosc_key)
                    row_items: List[QStandardItem] = [
                        QStandardItem(kierunek),
                        QStandardItem(specjalnosc),
                        QStandardItem(str(godziny["Zimowy stacjonarne"])),
                        QStandardItem(str(godziny["Zimowy niestacjonarne"])),
                        QStandardItem(str(godziny["Letni stacjonarne"])),
                        QStandardItem(str(godziny["Letni niestacjonarne"])),
                        QStandardItem(str(godziny["Suma"]))
                    ]
                    self.summary_model.appendRow(row_items)
                    # Dodaj do sum kierunku
                    suma_kierunku["Zimowy stacjonarne"] = float(suma_kierunku["Zimowy stacjonarne"]) + float(godziny["Zimowy stacjonarne"])
                    suma_kierunku["Zimowy niestacjonarne"] = float(suma_kierunku["Zimowy niestacjonarne"]) + float(godziny["Zimowy niestacjonarne"])
                    suma_kierunku["Letni stacjonarne"] = float(suma_kierunku["Letni stacjonarne"]) + float(godziny["Letni stacjonarne"])
                    suma_kierunku["Letni niestacjonarne"] = float(suma_kierunku["Letni niestacjonarne"]) + float(godziny["Letni niestacjonarne"])
                    suma_kierunku["Suma"] = float(suma_kierunku["Suma"]) + float(godziny["Suma"])
                # Dodaj wiersz sumujący dla kierunku
                row_items = [
                    QStandardItem(kierunek),
                    QStandardItem("SUMA kierunku"),
                    QStandardItem(str(suma_kierunku["Zimowy stacjonarne"])),
                    QStandardItem(str(suma_kierunku["Zimowy niestacjonarne"])),
                    QStandardItem(str(suma_kierunku["Letni stacjonarne"])),
                    QStandardItem(str(suma_kierunku["Letni niestacjonarne"])),
                    QStandardItem(str(suma_kierunku["Suma"]))
                ]
                self.summary_model.appendRow(row_items)

            if not kierunek_dict:
                self.summary_model.setHorizontalHeaderLabels(["Brak danych do wyświetlenia."])
            self.save_current_filtered_summary()
            if hasattr(self, 'chart_widget'):
                chart_data: List[Dict[str, Any]] = []
                for row_idx in range(self.summary_model.rowCount()):
                    row_data: Dict[str, Any] = {}
                    for col_idx in range(self.summary_model.columnCount()):
                        header: Any = self.summary_model.headerData(col_idx, Qt.Horizontal)
                        item: Optional[QStandardItem] = self.summary_model.item(row_idx, col_idx)
                        if item:
                            try:
                                # Spróbuj przekonwertować na liczbę jeśli to możliwe
                                value: Union[float, str] = float(item.text()) if item.text().replace('.', '').isdigit() else item.text()
                            except:
                                value = item.text()
                            row_data[str(header)] = value
                    chart_data.append(row_data)
                self.chart_widget.set_data(chart_data)

        except Exception as e:
            self.summary_model.setHorizontalHeaderLabels(["Błąd"])
            self.summary_model.appendRow([QStandardItem(f"Błąd: {str(e)}")])

    def display_employee_workload(self, item: QListWidgetItem) -> None:
        """Display workload data for the selected employee."""
        # This method seems to be for QListWidget, but the current implementation uses QTableView.
        # It's kept for completeness but might need adjustment based on actual UI.
        # Assuming self.instructor_details is a QListWidget for this method.
        if not hasattr(self, 'instructor_details') or not isinstance(self.instructor_details, QListWidget):
            print("Error: self.instructor_details is not a QListWidget or not initialized.")
            return

        self.instructor_details.clear()
        selected_employee_id: Optional[int] = item.data(1)
        selected_year: str = self.year_filter.currentText()
        selected_unit: Optional[int] = self.unit_filter.currentData()

        if not selected_employee_id:
            self.instructor_details.addItem("Nie wybrano wykładowcy.")
            return

        db: Session = SessionLocal()
        try:
            group_data: List[Dict[str, Any]]
            if self.chceckbox.isChecked():
                group_data = get_group_data(selected_year, selected_unit, selected_employee_id, self.current_filtered_groups)
            else:
                group_data = get_group_data(selected_year, selected_unit, selected_employee_id, None)

            filtered_groups: Optional[List[Dict[str, Any]]] = self.current_filtered_groups if self.chceckbox.isChecked() else None
            workload_data: Dict[str, Any] = calculate_workload_for_employee(selected_employee_id, selected_year, selected_unit, filtered_groups)

            # Wyświetl szczegóły obciążenia dydaktycznego
            self.instructor_details.addItem(f"Stanowisko: {workload_data['stanowisko']}")
            self.instructor_details.addItem(f"Pensum uczelniane: {workload_data['pensum_uczelniane']}")
            self.instructor_details.addItem(f"Umowa od: {workload_data['umowa_pocz']} do: {workload_data['umowa_kon']}")
            self.instructor_details.addItem(f"Pensum: {workload_data['pensum']}")
            self.instructor_details.addItem(f"Godziny dydaktyczne Z stacjonarne: {workload_data['godziny_dydaktyczne_z_stacjonarne']}")
            self.instructor_details.addItem(f"Godziny dydaktyczne Z niestacjonarne: {workload_data['godziny_dydaktyczne_z_niestacjonarne']}")
            self.instructor_details.addItem(f"Godziny dydaktyczne L stacjonarne: {workload_data['godziny_dydaktyczne_l_stacjonarne']}")
            self.instructor_details.addItem(f"Godziny dydaktyczne L niestacjonarne: {workload_data['godziny_dydaktyczne_l_niestacjonarne']}")
            self.instructor_details.addItem(f"Nadgodziny/Niedobór: {workload_data['nadgodziny']}")
            self.instructor_details.addItem(f"Łączna zniżka: {workload_data['zniżka']} godzin")
            self.instructor_details.addItem(f"Etat: {workload_data['etat']}")
            self.instructor_details.addItem(f"Czy podstawowe miejsce pracy: {workload_data['CZY_PODSTAWOWE']}")
            # Wyświetl szczegóły zniżek
            self.instructor_details.addItem("Zniżki:")
            if workload_data.get("typy_znizek"):
                for znizka, godziny in zip(workload_data["typy_znizek"], workload_data.get("godziny_znizek", [])):
                    self.instructor_details.addItem(f"  - Typ: {znizka}, Liczba godzin: {godziny}")
            else:
                self.instructor_details.addItem("  Brak zniżek")

            # Wyświetl szczegóły przedmiotów
            self.instructor_details.addItem("Przedmioty:")
            for group in group_data:
                self.instructor_details.addItem(
                    f"  - {group['Przedmiot']} ({group['Typ zajęć']}): {group['Liczba godzin']} godz. w {group['Semestr']} semestrze"
                )
        except Exception as e:
            self.instructor_details.addItem(f"Błąd: {str(e)}")
            print(f"Error: {str(e)}")
        finally:
            db.close()
    
    def select_columns_dialog(self, columns: List[str], title: str) -> Optional[List[str]]:
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Wybierz kolumny do eksportu - {title}")
        layout = QVBoxLayout(dialog)
        label = QLabel("Zaznacz kolumny do eksportu:")
        layout.addWidget(label)
        checkboxes: List[QCheckBox] = []
        for col in columns:
            cb = QCheckBox(col)
            cb.setChecked(True)
            layout.addWidget(cb)
            checkboxes.append(cb)
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        layout.addWidget(buttons)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        if dialog.exec_() == QDialog.Accepted:
            selected: List[str] = [cb.text() for cb in checkboxes if cb.isChecked()]
            return selected
        else:
            return None

    def generate_report_from_db(self) -> None:
        """Generate an Excel report with improved formatting, nagłówek i stopka, wybór kolumn."""
        db: Session = SessionLocal()
        selected_unit: Optional[int] = self.unit_filter.currentData()
        selected_year: str = self.year_filter.currentText()
        selected_employee: Optional[int] = self.employee_filter.currentData()
        try:
            self.populate_summary()
            # Query employees and filter by the selected unit
            query = (
                db.query(Employee, Person)
                .join(Person, Employee.OS_ID == Person.ID)
                .join(GroupInstructor, GroupInstructor.PRAC_ID == Employee.ID)
                .join(Group, GroupInstructor.ZAJ_CYK_ID == Group.ZAJ_CYK_ID)
                .join(DidacticCycleClasses, Group.ZAJ_CYK_ID == DidacticCycleClasses.ID)
                .join(DidacticCycles, DidacticCycleClasses.CDYD_KOD == DidacticCycles.KOD)
                .filter(DidacticCycles.OPIS.like(f"%{selected_year}%"))
            )
            if selected_unit:
                query = query.filter(GroupInstructor.JEDN_KOD == selected_unit)
            if selected_employee:
                query = query.filter(Employee.ID == selected_employee)
            employees: List[tuple[Employee, Person]] = query.all()
            lp: int = 1
            data: List[Dict[str, Any]] = []
            for employee, person in employees:
                person_obj: Optional[Person] = db.query(Person).filter_by(ID=person.ID).first()
                organizational_unit: Optional[OrganizationalUnits] = db.query(OrganizationalUnits).filter_by(KOD=person_obj.JED_ORG_KOD).first() if person_obj else None
                workload_data: Dict[str, Any] = calculate_workload_for_employee(employee.ID, selected_year, selected_unit)
                tytul: Optional[Title] = db.query(Title).filter_by(ID=person_obj.TYTUL_PRZED).first() if person_obj else None
                data.append({
                    "Lp.": lp,
                    "Tytuły": tytul.NAZWA if tytul else "N/A",
                    "Nazwisko i imię": f"{person_obj.NAZWISKO} {person_obj.IMIE}" if person_obj else "N/A",
                    "J.O.": organizational_unit.OPIS if organizational_unit else "N/A",
                    "Forma": "etat" if workload_data["umowa_pocz"] != "Brak daty rozpoczęcia umowy" else "umowa zlecenie",
                    "Stanowisko": workload_data["stanowisko"],
                    "Umowa od": workload_data["umowa_pocz"],
                    "Umowa do": workload_data["umowa_kon"],
                    "Pensum uczelniane": workload_data["pensum_uczelniane"],
                    "Zniżka": workload_data["zniżka"],
                    "Czy podstawowe miejsce pracy": workload_data["CZY_PODSTAWOWE"],
                    "Godziny dydaktyczne Z stacjonarne": workload_data["godziny_dydaktyczne_z_stacjonarne"],
                    "Godziny dydaktyczne Z niestacjonarne": workload_data["godziny_dydaktyczne_z_niestacjonarne"],
                    "Godziny dydaktyczne L stacjonarne": workload_data["godziny_dydaktyczne_l_stacjonarne"],
                    "Godziny dydaktyczne L niestacjonarne": workload_data["godziny_dydaktyczne_l_niestacjonarne"],
                    "Pensum realne": workload_data["total_workload"],
                    "Pensum": workload_data["pensum"],
                    "Etat": workload_data["etat"],
                    "Nadgodziny": workload_data["nadgodziny"],
                    "Stawka": workload_data["stawka"],
                    "Kwota nadgodzin": workload_data["kwota_nadgodzin"],
                })
                lp += 1
            # Wybór kolumn dla Wykładowców
            df1: pd.DataFrame
            if data:
                all_columns: List[str] = list(data[0].keys())
                selected_columns: Optional[List[str]] = self.select_columns_dialog(all_columns, "Wykładowcy")
                if not selected_columns:
                    self.status_label.setText("Status: Anulowano eksport.")
                    db.close()
                    return
                df1 = pd.DataFrame(data)[selected_columns]
            else:
                df1 = pd.DataFrame()
            # Grupy
            data2: List[Dict[str, Any]] = get_group_data(selected_year, selected_unit, selected_employee)
            df2: pd.DataFrame
            if data2:
                all_columns2: List[str] = list(data2[0].keys())
                selected_columns2: Optional[List[str]] = self.select_columns_dialog(all_columns2, "Grupy")
                if not selected_columns2:
                    self.status_label.setText("Status: Anulowano eksport.")
                    db.close()
                    return
                df2 = pd.DataFrame(data2)[selected_columns2]
            else:
                df2 = pd.DataFrame()
            # Podsumowanie (z sumą kierunku)
            summary_data: List[Dict[str, Any]] = []
            group_data_summary: List[Dict[str, Any]] = get_group_data(selected_year, selected_unit, selected_employee)
            kierunek_dict_summary: Dict[str, Dict[str, Dict[str, Union[int, float]]]] = {}
            for group in group_data_summary:
                kierunek: str = group.get("Kierunek", "Nieznany kierunek")
                specjalnosc: str = group.get("Specjalność", "Brak specjalności")
                tryb: str = group.get("Tryb", "Nieznany tryb").strip().lower()
                hours: Union[int, float] = group.get("Liczba godzin", 0)
                semester: str = group.get("Semestr", "Nieznany semestr").lower()
                if kierunek not in kierunek_dict_summary:
                    kierunek_dict_summary[kierunek] = {}
                if specjalnosc not in kierunek_dict_summary[kierunek]:
                    kierunek_dict_summary[kierunek][specjalnosc] = {
                        "Zimowy stacjonarne": 0,
                        "Zimowy niestacjonarne": 0,
                        "Letni stacjonarne": 0,
                        "Letni niestacjonarne": 0,
                        "Suma": 0
                    }
                if "zimowy" in semester and (("niestacjonarne" in tryb) or tryb == "none"):
                    kierunek_dict_summary[kierunek][specjalnosc]["Zimowy niestacjonarne"] = float(kierunek_dict_summary[kierunek][specjalnosc]["Zimowy niestacjonarne"]) + hours
                elif "zimowy" in semester and "stacjonarne" in tryb:
                    kierunek_dict_summary[kierunek][specjalnosc]["Zimowy stacjonarne"] = float(kierunek_dict_summary[kierunek][specjalnosc]["Zimowy stacjonarne"]) + hours
                elif "letni" in semester and (("niestacjonarne" in tryb) or tryb == "none"):
                    kierunek_dict_summary[kierunek][specjalnosc]["Letni niestacjonarne"] = float(kierunek_dict_summary[kierunek][specjalnosc]["Letni niestacjonarne"]) + hours
                elif "letni" in semester and "stacjonarne" in tryb:
                    kierunek_dict_summary[kierunek][specjalnosc]["Letni stacjonarne"] = float(kierunek_dict_summary[kierunek][specjalnosc]["Letni stacjonarne"]) + hours
                kierunek_dict_summary[kierunek][specjalnosc]["Suma"] = float(kierunek_dict_summary[kierunek][specjalnosc]["Suma"]) + hours
            for kierunek, specjalnosci in kierunek_dict_summary.items():
                suma_kierunku: Dict[str, Union[int, float]] = {
                    "Zimowy stacjonarne": 0,
                    "Zimowy niestacjonarne": 0,
                    "Letni stacjonarne": 0,
                    "Letni niestacjonarne": 0,
                    "Suma": 0
                }
                for specjalnosc, godziny in specjalnosci.items():
                    summary_data.append({
                        "Kierunek": kierunek,
                        "Specjalność": specjalnosc,
                        "Zimowy stacjonarne": godziny["Zimowy stacjonarne"],
                        "Zimowy niestacjonarne": godziny["Zimowy niestacjonarne"],
                        "Letni stacjonarne": godziny["Letni stacjonarne"],
                        "Letni niestacjonarne": godziny["Letni niestacjonarne"],
                        "Suma": godziny["Suma"]
                    })
                    suma_kierunku["Zimowy stacjonarne"] = float(suma_kierunku["Zimowy stacjonarne"]) + float(godziny["Zimowy stacjonarne"])
                    suma_kierunku["Zimowy niestacjonarne"] = float(suma_kierunku["Zimowy niestacjonarne"]) + float(godziny["Zimowy niestacjonarne"])
                    suma_kierunku["Letni stacjonarne"] = float(suma_kierunku["Letni stacjonarne"]) + float(godziny["Letni stacjonarne"])
                    suma_kierunku["Letni niestacjonarne"] = float(suma_kierunku["Letni niestacjonarne"]) + float(godziny["Letni niestacjonarne"])
                    suma_kierunku["Suma"] = float(suma_kierunku["Suma"]) + float(godziny["Suma"])
                summary_data.append({
                    "Kierunek": kierunek,
                    "Specjalność": "SUMA kierunku",
                    "Zimowy stacjonarne": suma_kierunku["Zimowy stacjonarne"],
                    "Zimowy niestacjonarne": suma_kierunku["Zimowy niestacjonarne"],
                    "Letni stacjonarne": suma_kierunku["Letni stacjonarne"],
                    "Letni niestacjonarne": suma_kierunku["Letni niestacjonarne"],
                    "Suma": suma_kierunku["Suma"]
                })
            df3: pd.DataFrame
            if summary_data:
                all_columns3: List[str] = list(summary_data[0].keys())
                selected_columns3: Optional[List[str]] = self.select_columns_dialog(all_columns3, "Podsumowanie")
                if not selected_columns3:
                    self.status_label.setText("Status: Anulowano eksport.")
                    db.close()
                    return
                else:
                    df3 = pd.DataFrame(summary_data)[selected_columns3]
            else:
                df3 = pd.DataFrame()
            now: datetime = datetime.now()
            default_name: str = f"raport_{now.strftime('%Y-%m-%d_%H-%M')}.xlsx"
            file_path: str
            file_path, _ = QFileDialog.getSaveFileName(self, "Save File", default_name, "Excel Files (*.xlsx)")
            if file_path:
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    # Eksport bez nagłówków tekstowych i bez nagłówków kolumn
                    if not df1.empty:
                        df1.to_excel(writer, sheet_name='Wykładowcy', index=False, header=True)
                    if not df2.empty:
                        df2.to_excel(writer, sheet_name='Grupy', index=False, header=True)
                    if not df3.empty:
                        df3.to_excel(writer, sheet_name='Podsumowanie', index=False, header=True)
                # Dodaj stopkę z datą do arkuszy
                self.add_footer_to_excel(file_path)
                self.add_pivot_table_to_excel(file_path, df2)
                self.format_excel(file_path)
                self.add_charts_to_excel(file_path)
                self.status_label.setText(f"Status: Raport zapisany do {file_path}")
            else:
                self.status_label.setText("Status: Anulowano zapis raportu")
        except Exception as e:
            self.status_label.setText(f"Status: Błąd podczas generowania raportu: {str(e)}")
        finally:
            db.close()

    def add_footer_to_excel(self, file_path: str) -> None:
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        date_str: str = datetime.now().strftime("Data wygenerowania raportu: %Y-%m-%d %H:%M")
        for sheet_name in ["Wykładowcy", "Grupy"]:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                last_row: int = ws.max_row + 2
                ws.cell(row=last_row, column=1, value=date_str)
        wb.save(file_path)

    def format_excel(self,file_path: str) -> None:
        """
        Apply formatting to the Excel file, adjust column widths, and add Excel tables with filtering.
        This version detects tables based on contiguous data blocks (separated by empty rows).
        """
        wb = load_workbook(file_path)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            # Step 1: Identify table ranges (contiguous blocks of data)
            table_ranges = self.find_table_ranges(sheet)

            # Step 2: Apply formatting and add tables for each identified range
            for i, (start_row, end_row, start_col, end_col) in enumerate(table_ranges):
                # Ensure there's at least one header row and some data
                if end_row <= start_row or end_col < start_col:
                    continue

                # Apply formatting to the header row of the current table
                for col_idx in range(start_col, end_col + 1):
                    cell = sheet.cell(row=start_row, column=col_idx)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border

                # Apply formatting to the rest of the cells in the current table
                for row_idx in range(start_row + 1, end_row + 1):
                    for col_idx in range(start_col, end_col + 1):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                        cell.border = thin_border

                # Adjust column widths for columns within the current table range
                # This part is tricky if tables share columns. For simplicity, we'll adjust
                # based on the content of the *entire* column, but only for columns
                # that are part of the current table.
                for col_idx in range(start_col, end_col + 1):
                    max_length: int = 0
                    column_letter: str = get_column_letter(col_idx)
                    # Iterate only through rows relevant to this table for width calculation
                    for row_idx in range(start_row, end_row + 1):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        try:
                            if cell.value is not None:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width: int = max_length + 2
                    # Only apply if the calculated width is greater than current to avoid shrinking
                    if sheet.column_dimensions[column_letter].width is None or adjusted_width > sheet.column_dimensions[column_letter].width:
                        sheet.column_dimensions[column_letter].width = adjusted_width

                # Add Excel table with filtering for the current table range
                table_ref: str = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
                
                # Generate a unique display name for each table
                table_display_name = f"Table_{sheet.title.replace(' ', '_')}_{i+1}"

                table = Table(displayName=table_display_name, ref=table_ref)
                style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                    showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                table.tableStyleInfo = style
                
                # Remove existing tables with the same name (if any)
                # This is important if you run the function multiple times on the same file
                if table.displayName in sheet.tables:
                    del sheet.tables[table.displayName]
                
                sheet.add_table(table)

        wb.save(file_path)

    def find_table_ranges(self, sheet: Any) -> List[Tuple[int, int, int, int]]:
        """
        Identifies contiguous blocks of data (tables) in a worksheet.
        A table is defined as a block of cells where there are no empty rows
        within the block, and it's surrounded by empty rows or sheet boundaries.
        Returns a list of tuples: (start_row, end_row, start_col, end_col) for each table.
        """
        table_ranges: List[Tuple[int, int, int, int]] = []
        in_table: bool = False
        current_table_start_row: int = -1
        current_table_end_row: int = -1
        
        # Iterate through rows to find table boundaries
        for row_idx in range(1, sheet.max_row + 1):
            # Check if the row is empty (all cells are None or empty string)
            is_row_empty = True
            for col_idx in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if cell_value is not None and str(cell_value).strip() != "":
                    is_row_empty = False
                    break
            
            if not in_table and not is_row_empty:
                # Start of a new table
                in_table = True
                current_table_start_row = row_idx
                current_table_end_row = row_idx
            elif in_table and not is_row_empty:
                # Continue current table
                current_table_end_row = row_idx
            elif in_table and is_row_empty:
                # End of current table
                # Now find the actual column range for this table
                start_col, end_col = self.find_column_range(sheet, current_table_start_row, current_table_end_row)
                if start_col != -1 and end_col != -1: # Ensure valid column range
                    table_ranges.append((current_table_start_row, current_table_end_row, start_col, end_col))
                in_table = False
                current_table_start_row = -1
                current_table_end_row = -1
        
        # Handle case where the last block of data is a table
        if in_table:
            start_col, end_col = self.find_column_range(sheet, current_table_start_row, current_table_end_row)
            if start_col != -1 and end_col != -1:
                table_ranges.append((current_table_start_row, current_table_end_row, start_col, end_col))
                
        return table_ranges

    def find_column_range(self, sheet: Any, start_row: int, end_row: int) -> Tuple[int, int]:
        """
        Finds the actual start and end column for a given table row range.
        """
        min_col = sheet.max_column + 1
        max_col = 0
        
        for row_idx in range(start_row, end_row + 1):
            for col_idx in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if cell_value is not None and str(cell_value).strip() != "":
                    min_col = min(min_col, col_idx)
                    max_col = max(max_col, col_idx)
                    
        if max_col == 0: # No data found in the given row range
            return -1, -1
        return min_col, max_col

    def filter_group_list(self, text: str) -> None:
        column: Optional[int] = self.group_filter_column_combo.currentData()
        if column is not None:
            self.group_filter_texts[column] = text
            self.group_proxy.setColumnFilter(column, text)
            self.update_group_active_filters()
            self.save_current_filtered_groups()
            if self.chceckbox.isChecked():
                self.changeFlag = True

    def filter_instructor_list(self, text: str) -> None:
        column: Optional[int] = self.instructor_filter_column_combo.currentData()
        if column is not None:
            self.instructor_filter_texts[column] = text
            self.instructor_proxy.setColumnFilter(column,text)
            self.update_instructor_active_filters()
            self.save_current_filtered_instructors()
            if self.chceckbox.isChecked():
                self.changeFlag = True
                # Jeśli po zmianie tekstu wszystkie filtry są puste, wyzwól odświeżenie przez przełączenie checkboxa
                if not any(v for v in self.instructor_filter_texts.values()):
                    #self.chceckbox.blockSignals(True)
                    self.chceckbox.setChecked(False)
                    self.chceckbox.setChecked(True)
                    #self.chceckbox.blockSignals(False)

    def filter_summary_list(self, text: str) -> None:
        column: Optional[int] = self.summary_filter_column_combo.currentData()
        if column is not None:
            self.summary_filter_texts[column] = text
            self.summary_proxy.setColumnFilter(column, text)
            self.update_summary_active_filters()
            self.save_current_filtered_summary()
            
            # Aktualizuj wykresy po filtrowaniu
            if hasattr(self, 'chart_widget'):
                filtered_data: List[Dict[str, Any]] = []
                for row_idx in range(self.summary_proxy.rowCount()):
                    row_data: Dict[str, Any] = {}
                    for col_idx in range(self.summary_proxy.columnCount()):
                        header: Any = self.summary_model.headerData(col_idx, Qt.Horizontal)
                        index: QModelIndex = self.summary_proxy.index(row_idx, col_idx)
                        value: Any = self.summary_proxy.data(index)
                        if value:
                            try:
                                value = float(value) if str(value).replace('.', '').isdigit() else value
                            except:
                                pass
                            row_data[str(header)] = value
                    filtered_data.append(row_data)
                
                self.chart_widget.set_data(filtered_data)
            
            if self.chceckbox.isChecked():
                self.changeFlag = True
    
    def on_group_filter_column_changed(self, index: int) -> None:
        column: Optional[int] = self.group_filter_column_combo.currentData()
        if column is not None:
            self.group_proxy.setFilterKeyColumn(column)
            self.group_search.blockSignals(True)
            self.group_search.setText(self.group_filter_texts.get(column, ""))
            self.group_search.blockSignals(False)
            self.update_group_active_filters()
            if self.chceckbox.isChecked():
                self.changeFlag = True

    def update_group_filter_columns(self, headers: List[str]) -> None:
        # Zapisz aktualny wybór
        current_data: Optional[int] = self.group_filter_column_combo.currentData()
        current_text: str = self.group_filter_column_combo.currentText()
        
        self.group_filter_column_combo.blockSignals(True)
        self.group_filter_column_combo.clear()
        
        for i, header in enumerate(headers):
            self.group_filter_column_combo.addItem(header, i)
        
        # Przywróć poprzedni wybór jeśli możliwe
        if current_data is not None:
            index: int = self.group_filter_column_combo.findData(current_data)
            if index != -1:
                self.group_filter_column_combo.setCurrentIndex(index)
        elif current_text:
            index = self.group_filter_column_combo.findText(current_text)
            if index != -1:
                self.group_filter_column_combo.setCurrentIndex(index)
        
        self.group_filter_column_combo.blockSignals(False)
    
    def on_instructor_filter_column_changed(self, index: int) -> None:
        column: Optional[int] = self.instructor_filter_column_combo.currentData()
        if column is not None:
            self.instructor_proxy.setFilterKeyColumn(column)
            self.instructor_search.blockSignals(True)
            self.instructor_search.setText(self.instructor_filter_texts.get(column, ""))
            self.instructor_search.blockSignals(False)
            self.update_instructor_active_filters()
            if self.chceckbox.isChecked():
                self.changeFlag = True

    def on_summary_filter_column_changed(self, index: int) -> None:
        column: Optional[int] = self.summary_filter_column_combo.currentData()
        if column is not None:
            self.summary_proxy.setFilterKeyColumn(column)
            self.summary_search.blockSignals(True)
            self.summary_search.setText(self.summary_filter_texts.get(column, ""))
            self.summary_search.blockSignals(False)
            self.update_summary_active_filters()
            if self.chceckbox.isChecked():
                self.changeFlag = True

    def update_instructor_filter_columns(self, headers: List[str]) -> None:
        # Zapisz aktualny wybór
        current_data: Optional[int] = self.instructor_filter_column_combo.currentData()
        current_text: str = self.instructor_filter_column_combo.currentText()
        
        self.instructor_filter_column_combo.blockSignals(True)
        self.instructor_filter_column_combo.clear()
        
        for i, header in enumerate(headers):
            self.instructor_filter_column_combo.addItem(header, i)
        
        # Przywróć poprzedni wybór jeśli możliwe
        if current_data is not None:
            index: int = self.instructor_filter_column_combo.findData(current_data)
            if index != -1:
                self.instructor_filter_column_combo.setCurrentIndex(index)
        elif current_text:
            index = self.instructor_filter_column_combo.findText(current_text)
            if index != -1:
                self.instructor_filter_column_combo.setCurrentIndex(index)
        
        self.instructor_filter_column_combo.blockSignals(False)

    def update_summary_filter_columns(self, headers: List[str]) -> None:
        # Zapisz aktualny wybór
        current_data: Optional[int] = self.summary_filter_column_combo.currentData()
        current_text: str = self.summary_filter_column_combo.currentText()
        
        self.summary_filter_column_combo.blockSignals(True)
        self.summary_filter_column_combo.clear()
        
        for i, header in enumerate(headers):
            self.summary_filter_column_combo.addItem(header, i)
        
        # Przywróć poprzedni wybór jeśli możliwe
        if current_data is not None:
            index: int = self.summary_filter_column_combo.findData(current_data)
            if index != -1:
                self.summary_filter_column_combo.setCurrentIndex(index)
        elif current_text:
            index = self.summary_filter_column_combo.findText(current_text)
            if index != -1:
                self.summary_filter_column_combo.setCurrentIndex(index)
        
        self.summary_filter_column_combo.blockSignals(False)
    
    def clear_group_filters(self) -> None:
        self.group_proxy.clearAllFilters()
        self.group_filter_texts.clear()
        self.group_search.blockSignals(True)
        self.group_search.clear()
        self.group_search.blockSignals(False)
        self.group_filter_column_combo.setCurrentIndex(0)
        self.update_group_active_filters()
        self.save_current_filtered_groups()
        if self.chceckbox.isChecked():
            #self.chceckbox.blockSignals(True)  # Blokujemy sygnały aby uniknąć rekurencji
            self.chceckbox.setChecked(False)
            self.chceckbox.setChecked(True)  # Automatycznie wywoła refresh_data()
            #self.chceckbox.blockSignals(False)
    
    def clear_instructor_filters(self) -> None:
        self.instructor_proxy.clearAllFilters()
        self.instructor_filter_texts.clear()
        self.instructor_search.blockSignals(True)
        self.instructor_search.clear()
        self.instructor_search.blockSignals(False)
        self.instructor_filter_column_combo.setCurrentIndex(0)
        self.update_instructor_active_filters()
        self.save_current_filtered_instructors()
        if self.chceckbox.isChecked():
            #self.chceckbox.blockSignals(True)  # Blokujemy sygnały aby uniknąć rekurencji
            self.chceckbox.setChecked(False)
            self.chceckbox.setChecked(True)  # Automatycznie wywoła refresh_data()
            #self.chceckbox.blockSignals(False)

    def clear_summary_filters(self) -> None:
        self.summary_proxy.clearAllFilters()
        self.summary_filter_texts.clear()
        self.summary_search.blockSignals(True)
        self.summary_search.clear()
        self.summary_search.blockSignals(False)
        self.summary_filter_column_combo.setCurrentIndex(0)
        self.update_summary_active_filters()
        self.save_current_filtered_summary()
        if self.chceckbox.isChecked():
            self.changeFlag = True

    def update_group_active_filters(self) -> None:
        # Usuń stare etykietki
        for i in reversed(range(self.group_active_filters_layout.count())):
            widget: Optional[QWidget] = self.group_active_filters_layout.itemAt(i).widget()
            if widget:
                widget.setParent(None)
        # Dodaj etykietki dla aktywnych filtrów
        for col, text in self.group_filter_texts.items():
            if text:
                idx: int = self.group_filter_column_combo.findData(col)
                col_name: str = self.group_filter_column_combo.itemText(idx) if idx != -1 else str(col)
                label = QLabel(f"{col_name}: {text}")
                btn = QPushButton("✖")
                btn.setObjectName("FilterRemove")
                btn.setFixedSize(24, 24)
                btn.clicked.connect(lambda _, c=col: self.remove_group_filter(c))
                filter_widget = QWidget()
                filter_layout = QHBoxLayout(filter_widget)
                filter_layout.setContentsMargins(2, 2, 2, 2)
                filter_layout.setSpacing(2)
                filter_layout.addWidget(label)
                filter_layout.addWidget(btn)
                self.group_active_filters_layout.addWidget(filter_widget)
                if self.chceckbox.isChecked():
                    self.changeFlag = True

    def remove_group_filter(self, column: int) -> None:
        self.group_filter_texts.pop(column, None)
        self.group_proxy.setColumnFilter(column, "")
        # Jeśli aktualnie wybrana kolumna, wyczyść pole wyszukiwania
        if self.group_filter_column_combo.currentData() == column:
            self.group_search.blockSignals(True)
            self.group_search.clear()
            self.group_search.blockSignals(False)
        self.update_group_active_filters()
        if self.chceckbox.isChecked():
            #self.chceckbox.blockSignals(True)  # Blokujemy sygnały aby uniknąć rekurencji
            self.chceckbox.setChecked(False)
            self.chceckbox.setChecked(True)
            #self.chceckbox.blockSignals(False)
    
    def update_instructor_active_filters(self) -> None:
        # Usuń stare etykietki
        for i in reversed(range(self.instructor_active_filters_layout.count())):
            widget: Optional[QWidget] = self.instructor_active_filters_layout.itemAt(i).widget()
            if widget:
                widget.setParent(None)
        # Dodaj etykietki dla aktywnych filtrów
        for col, text in self.instructor_filter_texts.items():
            if text:
                idx: int = self.instructor_filter_column_combo.findData(col)
                col_name: str = self.instructor_filter_column_combo.itemText(idx) if idx != -1 else str(col)
                label = QLabel(f"{col_name}: {text}")
                btn = QPushButton("✖")
                btn.setObjectName("FilterRemove")
                btn.setFixedSize(24, 24)
                btn.clicked.connect(lambda _, c=col: self.remove_instructor_filter(c))
                filter_widget = QWidget()
                filter_layout = QHBoxLayout(filter_widget)
                filter_layout.setContentsMargins(2, 2, 2, 2)
                filter_layout.setSpacing(2)
                filter_layout.addWidget(label)
                filter_layout.addWidget(btn)
                self.instructor_active_filters_layout.addWidget(filter_widget)
                if self.chceckbox.isChecked():
                    self.changeFlag = True
                
    def remove_instructor_filter(self, column: int) -> None:
        self.instructor_filter_texts.pop(column, None)
        self.instructor_proxy.setColumnFilter(column, "")
        # Jeśli aktualnie wybrana kolumna, wyczyść pole wyszukiwania
        if self.instructor_filter_column_combo.currentData() == column:
            self.instructor_search.blockSignals(True)
            self.instructor_search.clear()
            self.instructor_search.blockSignals(False)
        self.update_instructor_active_filters()
        if self.chceckbox.isChecked():
            # Jeśli po usunięciu filtra nie ma już żadnych aktywnych filtrów, przełącz checkbox (z sygnałami)
            if not any(v for v in self.instructor_filter_texts.values()):
                self.chceckbox.setChecked(False)
                self.chceckbox.setChecked(True)

    def update_summary_active_filters(self) -> None:
        # Usuń stare etykietki
        for i in reversed(range(self.summary_active_filters_layout.count())):
            widget: Optional[QWidget] = self.summary_active_filters_layout.itemAt(i).widget()
            if widget:
                widget.setParent(None)
        # Dodaj etykietki dla aktywnych filtrów
        for col, text in self.summary_filter_texts.items():
            if text:
                idx: int = self.summary_filter_column_combo.findData(col)
                col_name: str = self.summary_filter_column_combo.itemText(idx) if idx != -1 else str(col)
                label = QLabel(f"{col_name}: {text}")
                btn = QPushButton("✖")
                btn.setObjectName("FilterRemove")
                btn.setFixedSize(24, 24)
                btn.clicked.connect(lambda _, c=col: self.remove_summary_filter(c))
                filter_widget = QWidget()
                filter_layout = QHBoxLayout(filter_widget)
                filter_layout.setContentsMargins(2, 2, 2, 2)
                filter_layout.setSpacing(2)
                filter_layout.addWidget(label)
                filter_layout.addWidget(btn)
                self.summary_active_filters_layout.addWidget(filter_widget)
                if self.chceckbox.isChecked():
                    self.changeFlag = True
                
    def remove_summary_filter(self, column: int) -> None:
        self.summary_filter_texts.pop(column, None)
        self.summary_proxy.setColumnFilter(column, "")
        # Jeśli aktualnie wybrana kolumna, wyczyść pole wyszukiwania
        if self.summary_filter_column_combo.currentData() == column:
            self.summary_search.blockSignals(True)
            self.summary_search.clear()
            self.summary_search.blockSignals(False)
        self.update_summary_active_filters()
        if self.chceckbox.isChecked():
            self.changeFlag = True

    def generate_report_from_view(self) -> None:
        try:
            group_headers: List[str] = [str(self.group_model.headerData(i, Qt.Orientation.Horizontal)) for i in range(self.group_model.columnCount())]
            instructor_headers: List[str] = [str(self.instructor_model.headerData(i, Qt.Orientation.Horizontal)) for i in range(self.instructor_model.columnCount())]
            summary_headers: List[str] = [str(self.summary_model.headerData(i, Qt.Orientation.Horizontal)) for i in range(self.summary_model.columnCount())]

            group_data: List[Dict[str, Any]] = self.get_visible_table_data(self.group_proxy, group_headers)
            instructor_data: List[Dict[str, Any]] = self.get_visible_table_data(self.instructor_proxy, instructor_headers)
            summary_data: List[Dict[str, Any]] = self.get_visible_table_data(self.summary_proxy, summary_headers)
            
            df1: pd.DataFrame
            if instructor_data:
                all_columns: List[str] = list(instructor_data[0].keys())
                selected_columns: Optional[List[str]] = self.select_columns_dialog(all_columns, "Wykładowcy")
                if not selected_columns:
                    self.status_label.setText("Status: Anulowano zapis raportu")
                    return
                df1 = pd.DataFrame(instructor_data)[selected_columns]
            else:
                df1 = pd.DataFrame()
            
            df2: pd.DataFrame
            if group_data:
                all_columns_group: List[str] = list(group_data[0].keys())
                selected_columns_group: Optional[List[str]] = self.select_columns_dialog(all_columns_group, "Grupy")
                if not selected_columns_group:
                    self.status_label.setText("Status: Anulowano zapis raportu")
                    return
                df2 = pd.DataFrame(group_data)[selected_columns_group]
            else:
                df2 = pd.DataFrame()
            
            df3: pd.DataFrame
            if summary_data:
                all_columns_summary: List[str] = list(summary_data[0].keys())
                selected_columns_summary: Optional[List[str]] = self.select_columns_dialog(all_columns_summary, "Podsumowanie")
                if not selected_columns_summary:
                    self.status_label.setText("Status: Anulowano zapis raportu")
                    return
                df3 = pd.DataFrame(summary_data)[selected_columns_summary]
            else:
                df3 = pd.DataFrame()
            
            now: datetime = datetime.now()
            default_name: str = f"raport_{now.strftime('%Y-%m-%d_%H-%M')}.xlsx"
            file_path: str
            file_path, _ = QFileDialog.getSaveFileName(self, "Save File", default_name, "Excel Files (*.xlsx)")
            if file_path:
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    # Eksport bez nagłówków tekstowych
                    if not df1.empty:
                        df1.to_excel(writer, sheet_name='Wykładowcy', index=False)
                    if not df2.empty:
                        df2.to_excel(writer, sheet_name='Grupy', index=False)
                    if not df3.empty:
                        df3.to_excel(writer, sheet_name='Podsumowanie', index=False)
                self.add_footer_to_excel(file_path)
                self.add_pivot_table_to_excel(file_path, df2)
                self.format_excel(file_path)
                self.add_charts_to_excel(file_path)
                self.status_label.setText(f"Status: Raport zapisany do {file_path}")
            else:
                self.status_label.setText("Status: Anulowano zapis raportu")
        except Exception as e:
            print(e)
            self.status_label.setText(f"Status: Błąd podczas generowania raportu: {str(e)}")

    def add_charts_to_excel(self, file_path: str) -> None:
        """Generuje wszystkie wykresy i dodaje je do nowego arkusza w pliku Excel."""
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        # Utwórz nowy arkusz dla wykresów
        if "Wykresy" in wb.sheetnames:
            ws_charts = wb["Wykresy"]
        else:
            ws_charts = wb.create_sheet("Wykresy")
        # Ustaw dane dla wykresów (te same, co w zakładce podsumowania)
        chart_data: List[Dict[str, Any]] = []
        for row_idx in range(self.summary_proxy.rowCount()):
            row_data: Dict[str, Any] = {}
            for col_idx in range(self.summary_proxy.columnCount()):
                header: Any = self.summary_model.headerData(col_idx, Qt.Horizontal)
                item_index: QModelIndex = self.summary_proxy.index(row_idx, col_idx)
                value: Any = self.summary_proxy.data(item_index)
                if value:
                    try:
                        value = float(value) if str(value).replace('.', '').isdigit() else value
                    except:
                        pass
                    row_data[str(header)] = value
            chart_data.append(row_data)
        
        self.chart_widget.set_data(chart_data) # Upewnij się, że dane są ustawione w ChartWidget
        chart_types: List[str] = list(self.chart_widget.chart_buttons.keys())
        current_row: int = 1
        
        for chart_type in chart_types:
            try:
                self.chart_widget.update_chart(chart_type) # Wygeneruj wykres
                
                # Zapisz wykres do bufora jako obraz PNG
                img_buffer = BytesIO()
                self.chart_widget.figure.savefig(img_buffer, format='png', bbox_inches='tight', dpi=300)
                img_buffer.seek(0) # Przewiń bufor na początek
                
                # Utwórz obiekt obrazu openpyxl
                img = ExcelImage(img_buffer)
                
                # Ustaw rozmiar obrazu (opcjonalnie, dostosuj do potrzeb)
                #img.width = self.chart_widget.figure.get_figwidth() * self.chart_widget.figure.dpi
                #img.height = img.width
                
                # Dodaj obraz do arkusza
                cell_ref: str = f"A{current_row}"
                ws_charts.add_image(img, cell_ref)
                
                # Przesuń wiersz dla następnego wykresu (dodaj margines)
                current_row += int(img.height / 15) + 5 # Przybliżona wysokość wierszy
                
            except Exception as e:
                print(f"Błąd podczas eksportowania wykresu {chart_type}: {e}")
                ws_charts.cell(row=current_row, column=1, value=f"Błąd eksportu wykresu {chart_type}: {e}")
                current_row += 2 # Przesuń wiersz w przypadku błędu
        wb.save(file_path)
        print(f"Wykresy zostały dodane do pliku Excel: {file_path}")
    
    def get_visible_table_data(self, proxy_model: MultiColumnMultiValueFilterProxyModel, headers: List[str]) -> List[Dict[str, Any]]:
        data: List[Dict[str, Any]] = []
        for row_idx in range(proxy_model.rowCount()):
            row_data: Dict[str, Any] = {}
            for col_idx, header in enumerate(headers):
                index: QModelIndex = proxy_model.index(row_idx, col_idx)
                row_data[header] = proxy_model.data(index)
            data.append(row_data)
        return data
    
    def generate_report(self) -> None:
        msg = QMessageBox()
        msg.setWindowTitle("Wybór źródła danych")
        msg.setText("Wybierz źródło danych, z którego chcesz wygenerować raport:")
        widok_btn = msg.addButton("Z widoku (wyświetlane dane w tabelach)", QMessageBox.AcceptRole)
        baza_btn = msg.addButton("Z bazy (wszystkie dostępne dane)", QMessageBox.DestructiveRole)
        msg.exec_()

        if msg.clickedButton() == widok_btn:
            self.generate_report_from_view()
        elif msg.clickedButton() == baza_btn:
            self.generate_report_from_db()

    def toggle_theme(self) -> None:
        if self.is_dark_mode:
            self.setStyleSheet(light_stylesheet)
            self.is_dark_mode = False
            self.theme_toggle_btn.setText("🌜")
        else:
            self.setStyleSheet(dark_stylesheet)
            self.is_dark_mode = True
            self.theme_toggle_btn.setText("☀️")

    def save_current_filtered_groups(self) -> None:
        """Zapisuje aktualny stan przefiltrowanej tabeli grup"""
        self.current_filtered_groups = []
        headers: List[str] = [str(self.group_model.headerData(i, Qt.Orientation.Horizontal)) 
                for i in range(self.group_model.columnCount())]
        
        for row_idx in range(self.group_proxy.rowCount()):
            row_data: Dict[str, Any] = {}
            for col_idx, header in enumerate(headers):
                index: QModelIndex = self.group_proxy.index(row_idx, col_idx)
                row_data[header] = self.group_proxy.data(index)
            self.current_filtered_groups.append(row_data)
        
        # Opcjonalnie wyświetl informację o liczbie wierszy
        print(f"Zapisano stan tabeli grup: {len(self.current_filtered_groups)} wierszy")
        self.status_label.setText(f"Status: Przefiltrowano grupy - {len(self.current_filtered_groups)} wierszy")

    def save_current_filtered_instructors(self) -> None:
        """Zapisuje aktualny stan przefiltrowanej tabeli wykładowców"""
        self.current_filtered_instructors = []
        headers: List[str] = [str(self.instructor_model.headerData(i, Qt.Orientation.Horizontal)) 
                for i in range(self.instructor_model.columnCount())]
        
        for row_idx in range(self.instructor_proxy.rowCount()):
            row_data: Dict[str, Any] = {}
            for col_idx, header in enumerate(headers):
                index: QModelIndex = self.instructor_proxy.index(row_idx, col_idx)
                row_data[header] = self.instructor_proxy.data(index)
            self.current_filtered_instructors.append(row_data)
        
        print(f"Zapisano stan tabeli wykładowców: {len(self.current_filtered_instructors)} wierszy")
        self.status_label.setText(f"Status: Przefiltrowano wykładowców - {len(self.current_filtered_instructors)} wierszy")

    def save_current_filtered_summary(self) -> None:
        """Zapisuje aktualny stan przefiltrowanej tabeli podsumowania"""
        self.current_filtered_summary = []
        headers: List[str] = [str(self.summary_model.headerData(i, Qt.Orientation.Horizontal)) 
                for i in range(self.summary_model.columnCount())]
        
        for row_idx in range(self.summary_proxy.rowCount()):
            row_data: Dict[str, Any] = {}
            for col_idx, header in enumerate(headers):
                index: QModelIndex = self.summary_proxy.index(row_idx, col_idx)
                row_data[header] = self.summary_proxy.data(index)
            self.current_filtered_summary.append(row_data)
        
        print(f"Zapisano stan tabeli podsumowania: {len(self.current_filtered_summary)} wierszy")
        self.status_label.setText(f"Status: Przefiltrowano podsumowanie - {len(self.current_filtered_summary)} wierszy")
    
    def get_instructors_id(self, instructor_proxy: MultiColumnMultiValueFilterProxyModel) -> List[int]:
        """Zwraca listę ID wykładowców z aktualnie przefiltrowanej tabeli"""
        db: Session = SessionLocal()
        try:
            print(f"Pobieranie ID wykładowców z tabeli: {instructor_proxy.rowCount()} wierszy")
            # Znajdź indeks kolumny po nagłówku "Nazwisko i imię"
            name_col: Optional[int] = None
            for col in range(self.instructor_model.columnCount()):
                header: Any = self.instructor_model.headerData(col, Qt.Horizontal)
                if str(header) == "Nazwisko i imię":
                    name_col = col
                    break

            if name_col is None:
                print("Nie znaleziono kolumny 'Nazwisko i imię' w modelu wykładowców.")
                return []

            instructor_ids: List[int] = []
            for row_idx in range(instructor_proxy.rowCount()):
                index: QModelIndex = instructor_proxy.index(row_idx, name_col)
                name: Any = instructor_proxy.data(index)

                if name:
                    # Dopasowanie po pełnym "Nazwisko i imię" tak jak w display_instructor_details
                    try:
                        employee_obj: Optional[Employee] = (
                            db.query(Employee)
                            .join(Person, Person.ID == Employee.OS_ID)
                            .filter((Person.NAZWISKO + " " + Person.IMIE) == str(name))
                            .first()
                        )
                        if employee_obj:
                            instructor_ids.append(employee_obj.ID)
                        else:
                            print(f"Nie znaleziono osoby dla: {name}")
                    except Exception as e:
                        print(f"Błąd dopasowania osoby '{name}': {e}")
                else:
                    print("Puste imię i nazwisko w wierszu proxy.")

            print(f"Znalezione ID wykładowców: {instructor_ids}")
            return instructor_ids
        except Exception as e:
            print(f"Błąd podczas pobierania ID wykładowców: {str(e)}")
            return []
        finally:
            db.close()
    def add_pivot_table_to_excel(self, file_path: str, group_data_df: pd.DataFrame) -> None:
        """Dodaje arkusz z tabelą przestawną z formułami obliczającymi sumy."""
        
        try:
            # Sprawdź czy wymagane kolumny istnieją w danych
            required_columns = ['Prowadzący', 'Przedmiot', 'Semestr', 'Kierunek', 'Specjalność', 'Tryb', 'Stopień', 'Rok', 
                            'Instytut w którym jest rozliczany przedmiot', 
                            'Typ zajęć', 'Liczba godzin', 'Kod przedmiotu']
            
            # Sprawdź które kolumny są dostępne
            available_columns = [col for col in required_columns if col in group_data_df.columns]
            
            if len(available_columns) < len(required_columns):
                print(f"Brakujące kolumny w danych. Dostępne: {available_columns}")
                return
            
            # Utwórz kolumnę z liczbą grup (domyślnie 1)
            group_data_df['Liczba grup'] = 1
            
            # Utwórz tabele przestawne
            pivot_hours = group_data_df.pivot_table(
                values='Liczba godzin',
                index=['Prowadzący', 'Przedmiot', 'Semestr','Kierunek', 'Specjalność', 'Tryb', 'Stopień', 'Rok', 'Instytut w którym jest rozliczany przedmiot', 'Kod przedmiotu'],
                columns=['Typ zajęć'],
                aggfunc='first',  # Bierze pierwszą wartość (godziny dla jednej grupy)
                fill_value=0
            )
            
            pivot_groups = group_data_df.pivot_table(
                values='Liczba grup',
                index=['Prowadzący', 'Przedmiot', 'Semestr', 'Kierunek', 'Specjalność', 'Tryb', 'Stopień', 'Rok', 'Instytut w którym jest rozliczany przedmiot', 'Kod przedmiotu'],
                columns=['Typ zajęć'],
                aggfunc='sum',
                fill_value=0
            )
            
            # Przygotuj finalny DataFrame
            final_df = pivot_hours.reset_index()
            
            # Załaduj istniejący workbook
            wb = load_workbook(file_path)
            
            # Utwórz nowy arkusz
            if "Grupy planowanie" in wb.sheetnames:
                ws_pivot = wb["Grupy planowanie"]
                ws_pivot.delete_rows(1, ws_pivot.max_row)
            else:
                ws_pivot = wb.create_sheet("Grupy planowanie")
            
            # Zapisz podstawowe dane (tylko wiersze i kolumny podstawowe)
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            base_columns = ['Prowadzący', 'Przedmiot', 'Semestr', 'Kierunek', 'Specjalność', 'Tryb', 'Stopień', 'Rok', 'Instytut w którym jest rozliczany przedmiot']
            base_data = final_df[base_columns]
            
            for r_idx, row in enumerate(dataframe_to_rows(base_data, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws_pivot.cell(row=r_idx, column=c_idx, value=value)
            
            # Dodaj nagłówki dla każdego typu zajęć
            col_idx = len(base_columns) + 1
            zajecia_types = pivot_hours.columns
            
            for zajecia_type in zajecia_types:
                # Nagłówki kolumn dla tego typu zajęć
                ws_pivot.cell(row=1, column=col_idx, value=f"{zajecia_type} - Godziny")
                ws_pivot.cell(row=1, column=col_idx + 1, value=f"{zajecia_type} - Grupy")
                ws_pivot.cell(row=1, column=col_idx + 2, value=f"{zajecia_type} - Suma")
                col_idx += 3
            
            # Dodaj nagłówek dla sumy godzin na końcu wiersza
            ws_pivot.cell(row=1, column=col_idx, value="SUMA GODZIN")
            
            # Wypełnij dane i dodaj formuły
            for row_idx in range(2, len(final_df) + 2):  # +2 bo row 1 to nagłówek, row 2 to pierwsze dane
                col_idx = len(base_columns) + 1
                suma_formula_parts = []  # Przechowuje części formuły sumy
                
                for zajecia_type in zajecia_types:
                    # Godziny dla jednej grupy
                    hours_value = pivot_hours[zajecia_type].iloc[row_idx - 2]
                    ws_pivot.cell(row=row_idx, column=col_idx, value=hours_value)
                    
                    # Liczba grup
                    groups_value = pivot_groups[zajecia_type].iloc[row_idx - 2]
                    ws_pivot.cell(row=row_idx, column=col_idx + 1, value=groups_value)
                    
                    # Formuła suma = godziny * grupy
                    hours_cell = get_column_letter(col_idx) + str(row_idx)
                    groups_cell = get_column_letter(col_idx + 1) + str(row_idx)
                    formula = f"={hours_cell}*{groups_cell}"
                    ws_pivot.cell(row=row_idx, column=col_idx + 2, value=formula)
                    
                    # Dodaj komórkę sumy do formuły ogólnej sumy
                    suma_cell = get_column_letter(col_idx + 2) + str(row_idx)
                    suma_formula_parts.append(suma_cell)
                    
                    col_idx += 3
                
                # Dodaj formułę sumy na końcu wiersza
                if suma_formula_parts:
                    suma_formula = "=" + "+".join(suma_formula_parts)
                    ws_pivot.cell(row=row_idx, column=col_idx, value=suma_formula)
            
            # Dodaj wiersz z sumami
            last_row = len(final_df) + 3
            ws_pivot.cell(row=last_row, column=1, value="RAZEM")
            
            # Sumy dla każdej kolumny
            col_idx = len(base_columns) + 1
            
            for zajecia_type in zajecia_types:
                # Suma godzin
                hours_col = get_column_letter(col_idx)
                hours_sum_formula = f"=SUBTOTAL(9,{hours_col}2:{hours_col}{last_row - 1})"
                ws_pivot.cell(row=last_row, column=col_idx, value=hours_sum_formula)
                
                # Suma grup
                groups_col = get_column_letter(col_idx + 1)
                groups_sum_formula = f"=SUBTOTAL(9,{groups_col}2:{groups_col}{last_row - 1})"
                ws_pivot.cell(row=last_row, column=col_idx + 1, value=groups_sum_formula)
                
                # Suma sum (godziny * grupy)
                suma_col = get_column_letter(col_idx + 2)
                suma_sum_formula = f"=SUBTOTAL(9,{suma_col}2:{suma_col}{last_row - 1})"
                ws_pivot.cell(row=last_row, column=col_idx + 2, value=suma_sum_formula)
                
                col_idx += 3
            
            # Suma ogólna na końcu wiersza z sumami
            if zajecia_types.any():
                suma_total_col = get_column_letter(col_idx)
                suma_total_formula = f"=SUBTOTAL(9,{suma_total_col}2:{suma_total_col}{last_row - 1})"
                ws_pivot.cell(row=last_row, column=col_idx, value=suma_total_formula)
            
            # Dostosuj szerokości kolumn
            for column in ws_pivot.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 15)
                ws_pivot.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(file_path)
            print("Dodano tabelę przestawną z formułami do raportu Excel")
            
        except Exception as e:
            print(f"Błąd podczas tworzenia tabeli przestawnej: {str(e)}")
            import traceback
            traceback.print_exc()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    login_window = LoginWindow()
    login_window.show()
    sys.exit(app.exec_())
